from datetime import datetime, timedelta
from io import BytesIO
import pandas as pd


def export_detailed_timetable_to_excel(
        result: dict,
        dataset: pd.DataFrame,
        start_date: str,
        file_path: str | None = None,
        course_col: str = "course_id",
        room_col: str = "room",  # z.B. "Raum"
        instructor_col: str = "lecturer",  # z.B. "Dozent"
        title_col: str | None = None,  # z.B. "title" oder "modul"
        use_german_headers: bool = True
):
    # --- Prüfen, ob notwendige Spalten existieren ---
    required = [course_col]
    for opt in [room_col, instructor_col]:
        if opt is not None:
            required.append(opt)
    missing = [c for c in required if c not in dataset.columns]
    if missing:
        raise ValueError(f"Spalten im dataset fehlen: {missing}. "
                         f"Passe room_col/instructor_col/course_col an.")

    # --- Basis aus result ---
    weeks = int(result["weeks"])
    course_to_slot = result["course_to_slot"]  # {'K1': (W, 'Mo', 'Morning'), ...}

    # --- Startdatum -> Montag der Woche ---
    dt = datetime.strptime(start_date, "%Y-%m-%d").date()
    monday0 = dt - timedelta(days=dt.weekday())

    # --- Tages-/Beschriftungen ---
    day_order = ["Mo", "Di", "Mi", "Do", "Fr"]
    if use_german_headers:
        day_name = {"Mo": "Montag", "Di": "Dienstag", "Mi": "Mittwoch", "Do": "Donnerstag", "Fr": "Freitag"}
    else:
        day_name = {"Mo": "Monday", "Di": "Tuesday", "Mi": "Wednesday", "Do": "Thursday", "Fr": "Friday"}
    half_order = ["Morning", "Afternoon"]

    # --- Datensatz schlank mappen: course_id -> (Titel, Raum, Dozent)
    # Falls Mehrfachzeilen je Kurs existieren, nimm die erste (oder aggregiere nach Bedarf).
    cols_to_take = [course_col]
    for col in [title_col, room_col, instructor_col]:
        if col and col not in cols_to_take:
            cols_to_take.append(col)
    slim = dataset[cols_to_take].drop_duplicates(subset=[course_col]).set_index(course_col)

    def format_entry(cid: str):
        # hole Felder (mit Fallback auf "")
        title = slim[title_col].get(cid, "") if title_col else ""
        room = slim[room_col].get(cid, "") if room_col in slim.columns else ""
        inst = slim[instructor_col].get(cid, "") if instructor_col in slim.columns else ""
        parts = []
        # Kurs-ID immer zeigen
        parts.append(str(cid) if not title else f"{cid} — {title}")
        if room:
            parts.append(f"Raum: {room}")
        if inst:
            parts.append(f"Dozent: {inst}")
        return " — ".join(parts)

    # --- pro Woche eine Struktur (Tag, Half) -> Liste von format_entry
    per_week = {w: {(d, h): [] for d in day_order for h in half_order}
                for w in range(1, weeks + 1)}

    # füllen
    for cid, (w, d, h) in course_to_slot.items():
        if w in per_week and (d, h) in per_week[w]:
            per_week[w][(d, h)].append(format_entry(str(cid)))

    # --- Excel schreiben (openpyxl-Engine, keine zusätzliche Lib nötig) ---
    out = None
    if file_path is None:
        out = BytesIO()
        writer = pd.ExcelWriter(out, engine="openpyxl")
    else:
        writer = pd.ExcelWriter(file_path, engine="openpyxl")

    wb = writer.book

    # Stile (openpyxl)
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    fill_head = PatternFill("solid", fgColor="C8102E")  # DHBW-Rot
    fill_half = PatternFill("solid", fgColor="F2F2F2")
    font_white_bold = Font(color="FFFFFF", bold=True, size=12)
    font_black_bold = Font(color="000000", bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # --- Übersicht ---
    summary_df = pd.DataFrame({
        "Kennzahl": ["Startdatum (Woche 1, Montag)", "Wochen", "Kurse gesamt"],
        "Wert": [monday0.strftime("%d.%m.%Y"), weeks, len(course_to_slot)]
    })
    summary_df.to_excel(writer, index=False, sheet_name="Übersicht")
    ws_sum = writer.sheets["Übersicht"]
    ws_sum.column_dimensions["A"].width = 32
    ws_sum.column_dimensions["B"].width = 22
    ws_sum["A1"].font = Font(bold=True)

    # --- Wochenblätter ---
    for w in range(1, weeks + 1):
        ws_name = f"Woche {w}"
        # Dummy-DF, damit das Sheet existiert
        pd.DataFrame({"_": []}).to_excel(writer, index=False, sheet_name=ws_name)
        ws = writer.sheets[ws_name]
        # alles löschen
        ws.delete_rows(1, ws.max_row)

        # Spaltenbreiten
        ws.column_dimensions["A"].width = 16  # Slot
        for col_letter in ["B", "C", "D", "E", "F"]:
            ws.column_dimensions[col_letter].width = 36

        # Kopfzeile: Slot + Tage + Datum
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        cell = ws.cell(row=1, column=1, value="Slot")
        cell.fill = fill_head
        cell.font = font_white_bold
        cell.alignment = align_center
        cell.border = border

        week_monday = monday0 + timedelta(days=7 * (w - 1))
        # Zeile 1: Tagesnamen, Zeile 2: Datum
        for i, d in enumerate(day_order):
            col = 2 + i
            c1 = ws.cell(row=1, column=col, value=day_name[d])
            c1.fill = fill_head
            c1.font = font_white_bold
            c1.alignment = align_center
            c1.border = border

            the_date = (week_monday + timedelta(days=i)).strftime("%d.%m.%Y")
            c2 = ws.cell(row=2, column=col, value=the_date)
            c2.fill = fill_head
            c2.font = font_white_bold
            c2.alignment = align_center
            c2.border = border

        # Zeilenbeschriftungen Morning/Afternoon
        for r, half in enumerate(half_order, start=3):
            c = ws.cell(row=r, column=1, value=half)
            c.fill = fill_half
            c.font = font_black_bold
            c.alignment = align_center
            c.border = border

        # Zellen mit Kurslisten füllen
        for i, d in enumerate(day_order):
            col = 2 + i
            # Morning
            entries_m = per_week[w][(d, "Morning")]
            text_m = "\n".join(entries_m) if entries_m else ""
            cm = ws.cell(row=3, column=col, value=text_m)
            cm.alignment = align_center
            cm.border = border

            # Afternoon
            entries_a = per_week[w][(d, "Afternoon")]
            text_a = "\n".join(entries_a) if entries_a else ""
            ca = ws.cell(row=4, column=col, value=text_a)
            ca.alignment = align_center
            ca.border = border

        # Zeilenhöhen (mehr Platz)
        ws.row_dimensions[1].height = 26
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 90
        ws.row_dimensions[4].height = 90

    writer.close()

    if out is not None:
        out.seek(0)
        return out.getvalue()
    return None
